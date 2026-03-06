
from scripts.conf_file_finding import try_find_conf_file
try_find_conf_file()

import datajoint as dj
import pandas as pd
import time
from zoneinfo import ZoneInfo
import datetime
import pathlib
import numpy as np
import time
import base64
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

time.sleep(1)

import u19_pipeline.alert_system.water_weigh_alert.water_weigh_alert as wwa
import u19_pipeline.scheduler as scheduler
import u19_pipeline.acquisition as acquisition
import u19_pipeline.behavior as behavior
import u19_pipeline.subject as subject
import u19_pipeline.lab as lab


DJ_CUSTOM_VARIABLES_FILENAME = 'DJCustomVariables.csv'
SLACK_WEBHOOK_FILENAME = 'SlackChannels.csv'
USER_SLACK_FILENAME = 'UserSlack.csv'
RIG_STATUS_FILENAME = 'RigStatusTable.csv'
DAY_SCHEDULE_FILENAME = 'ScheduleDay.csv'
PAST_SESSION_PERFORMANCE_FILENAME = 'PastSessions.csv'

WEIGHING_GUI_REPLACEMENT_SPREADHSHEET_FILENAME_TEMPLATE = 'Weighing_GUI_Replacement_SpreadSheet_Template.xlsx'
WEIGHING_GUI_REPLACEMENT_SPREADHSHEET_FILENAME = 'Weighing_GUI_Replacement_SpreadSheet.xlsx'

conf = dj.config
nodb_virmen_backup_dir = pathlib.Path(pathlib.Path(conf['custom']['root_data_dir'][0]).parent.parent,'Shared','NoDBVirmenBackup')

MAX_SESSIONS_HISTORY = 75

def cast_choice(choice_array):

    new_array = choice_array[0].copy()
    new_array[new_array>2] = 127

    new_array = np.array(new_array, dtype=np.uint8)

    return new_array


def encode_webhook(webhook):

    return (base64.b64encode(webhook.encode('utf-8'))).decode('utf-8')
   


def write_dj_custom_vars_file():

    file_write = pathlib.Path(nodb_virmen_backup_dir, DJ_CUSTOM_VARIABLES_FILENAME)
    pd.DataFrame(lab.DjCustomVariables.fetch(as_dict=True)).to_csv(file_write)

def write_slack_webhooks_file():

    slack_webhooks = pd.DataFrame(lab.SlackWebhooks.fetch(as_dict=True))
    slack_webhooks['webhook_url'] = slack_webhooks['webhook_url'].astype(str)

    slack_webhooks['webhook_url'] = slack_webhooks['webhook_url'].apply(encode_webhook)

    file_write = pathlib.Path(nodb_virmen_backup_dir, SLACK_WEBHOOK_FILENAME)
    slack_webhooks.to_csv(file_write, index=False)

def write_user_data_file():


    user_data = pd.DataFrame(lab.User.fetch('user_id', 'slack', 'tech_responsibility', 'slack_webhook',as_dict=True))
    user_data['slack_webhook'] = user_data['slack_webhook'].astype(str)
    user_data['slack_webhook'] = user_data['slack_webhook'].fillna('')

    user_data['slack_webhook'] = user_data['slack_webhook'].apply(encode_webhook)

    file_write = pathlib.Path(nodb_virmen_backup_dir, USER_SLACK_FILENAME)
    user_data.to_csv(file_write, index=False)

def write_rig_status_file():

    file_write = pathlib.Path(nodb_virmen_backup_dir, RIG_STATUS_FILENAME)

    df_rig_status = pd.DataFrame(scheduler.RigStatus.fetch('location', 'input_output_name','current_status',as_dict=True))
    df_rig_status.to_csv(file_write, index=False)


def write_schedule_file():

    schedule_query = dict()
    schedule_query['date'] = datetime.date.today()

    subject_query = 'subject_fullname is not null'

    day_schedule = pd.DataFrame((scheduler.Schedule * scheduler.TrainingProfile * subject.Subject.proj(subject_user_id='user_id') & schedule_query & subject_query).fetch(as_dict=True))
    day_schedule = day_schedule.drop(columns='user_id')
    day_schedule = day_schedule.rename(columns={'subject_user_id':'user_id'})

    file_write = pathlib.Path(nodb_virmen_backup_dir, DAY_SCHEDULE_FILENAME)
    day_schedule.to_csv(file_write, index=False)

    return day_schedule


def write_past_sessions_file(day_schedule):

    all_subjects_schedule = "', '".join(day_schedule['subject_fullname'])
    all_subjects_schedule = "subject_fullname in ('" +all_subjects_schedule+ "')"
    
    ss = pd.DataFrame((behavior.TowersSession & all_subjects_schedule).fetch('KEY', order_by='subject_fullname, session_date desc, session_number desc', as_dict=True))
    ss['session_date'] = ss['session_date'].astype(str)
    ss['num_sessions'] = ss.groupby(['subject_fullname'])['subject_fullname'].rank(method='first')
    ss['num_sessions'] = ss['num_sessions'].astype(int)
    ss = ss.loc[ss['num_sessions'] <= MAX_SESSIONS_HISTORY, :]

    max_value_indices = ss.groupby('subject_fullname')['num_sessions'].idxmax()
    ss = ss.loc[max_value_indices]
    ss = ss.reset_index(drop=True)
    ss['query'] = "(subject_fullname='" + ss['subject_fullname'] + "' and session_date >= '" + ss['session_date'] + "')"

    block_query = ' OR '.join(ss['query'])

    sstable=(acquisition.SessionStarted.proj('local_path_behavior_file', 'session_location'))
    stable = (acquisition.Session).proj(stimulusBank='stimulus_bank')
    tstable = (behavior.TowersSession.proj(trialType='rewarded_side',choice='chosen_side', stimulusSet='stimulus_set'))  
    tbtable = (behavior.TowersBlock.proj('first_trial','n_trials', 'sublevel', mazeID='level', mainMazeID='main_level', easyBlockFlag='easy_block',\
                                        duration='block_duration', rewardMil='reward_mil', medianTrialDur='trial_duration_median', start='block_start_time'))  
    table_fetch = sstable * stable* tstable * tbtable

    allblocks = pd.DataFrame((table_fetch & block_query).fetch(order_by='subject_fullname, session_date desc, session_number desc, block desc',as_dict=True))
    allblocks['from_DB'] = 1
    allblocks['session_date'] = allblocks['session_date'].astype(str)

    allblocks['sublevel'] = allblocks['sublevel'].astype('Int64')
    allblocks['choice'] = allblocks['choice'].apply(cast_choice)
    allblocks['trialType'] = allblocks['trialType'].apply(cast_choice)

    file_write = pathlib.Path(nodb_virmen_backup_dir, PAST_SESSION_PERFORMANCE_FILENAME)
    allblocks.to_csv(file_write, index=False)


def write_weighinig_gui_ss_file():

    subject_data = wwa.get_subject_data()
    subject_data = subject_data.loc[:, ['user_id', 'subject_fullname', 'cage', 'headplate_image_path', 'last_weight', 'water_per_day']]
    #subject_data = subject_data.loc[subject_data['user_id'] != 'testuser',:]
    subject_data = subject_data.reset_index(drop=True)

    dictionary_rename_cols =\
    {'user_id': 'owner',
    'subject_fullname': 'subject name',
    'cage': 'cage',
    'headplate_image_path': 'headplate',
    'last_weight': 'last weight',
    'water_per_day': 'water per day',
    }

    subject_data = subject_data.rename(columns=dictionary_rename_cols)
    subject_data = subject_data[list(dictionary_rename_cols.values())]
    subject_data['headplate'] = subject_data['headplate'].astype(str)

    headplate_image_paths = subject_data['headplate'].copy()

    subject_data['last weight'] = subject_data['last weight'].round(1)
    subject_data['water per day'] = subject_data['water per day'].round(1)


    subject_data['headplate'] = ''
    subject_data['today weight'] = ''
    subject_data['today water'] = ''

    # Copy the source file to the destination, replacing if it exists
    template_ss_file = pathlib.Path(nodb_virmen_backup_dir,WEIGHING_GUI_REPLACEMENT_SPREADHSHEET_FILENAME_TEMPLATE)
    ss_file = pathlib.Path(nodb_virmen_backup_dir,WEIGHING_GUI_REPLACEMENT_SPREADHSHEET_FILENAME)
    shutil.copy2(template_ss_file, ss_file)

    book = openpyxl.load_workbook(ss_file)
    sheet = book['Sheet1']


    for r_idx, row in enumerate(dataframe_to_rows(subject_data, index=False, header=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)


    column_headplate = subject_data.columns.get_loc('headplate') + 1
    column_headplate = openpyxl.utils.get_column_letter(column_headplate)


    for i in range(headplate_image_paths.shape[0]):

        headplate_path = headplate_image_paths[i]
        cell_image = column_headplate+str(i+3)

        if pathlib.Path.is_file(pathlib.Path(headplate_path)):
            img = Image(headplate_path)
            img.width = 60
            img.height = 60
            sheet.add_image(img, cell_image)
            sheet.row_dimensions[i+3].height = 50 
        else:
            sheet.row_dimensions[i+3].height = 15

    book.save(ss_file)


def main_noDB_backup():

    write_dj_custom_vars_file()
    write_slack_webhooks_file()
    write_user_data_file()
    write_rig_status_file()
    day_schedule = write_schedule_file()
    write_past_sessions_file(day_schedule)
    write_weighinig_gui_ss_file()


if __name__ == "__main__":
    main_noDB_backup()

